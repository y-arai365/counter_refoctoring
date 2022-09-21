import openpyxl as px


def output_excel(load_path, save_path, first_column, control_no, figure_no, product_name, results):
    """

    :param load_path: 数量カウント.xlsxのパス
    :param save_path: 保存名
    :param first_column: 書き込み始める列(右側と左側があるため)
    :param control_no: 管理No
    :param figure_no: 製品図番
    :param product_name: 製品名
    :param results: 書き込む結果
    :return: エラーが発生しなければTrue, 発生すればFalse
    """
    try:
        # ワークシートを読み込む
        wb = px.load_workbook(load_path)
        ws = wb.active  # 編集可能にする

        # ロットの情報を入力する
        ws.unmerge_cells(start_row=3, start_column=first_column+3, end_row=3, end_column=first_column+5)  # セルの結合を解く
        ws.cell(row=3, column=first_column+3).value = control_no  # セルに値を入力する
        ws.merge_cells(start_row=3, start_column=first_column + 3, end_row=3, end_column=first_column + 5)  # セルを結合する
        ws.unmerge_cells(start_row=4, start_column=first_column + 3, end_row=4, end_column=first_column + 5)
        ws.cell(row=4, column=first_column+3).value = figure_no
        ws.merge_cells(start_row=4, start_column=first_column + 3, end_row=4, end_column=first_column + 5)
        ws.unmerge_cells(start_row=5, start_column=first_column + 3, end_row=5, end_column=first_column + 5)
        ws.cell(row=5, column=first_column+3).value = product_name
        ws.merge_cells(start_row=5, start_column=first_column + 3, end_row=5, end_column=first_column + 5)

        # カウント結果を入力する
        row = 9  # 書き込みはじめは9行目
        column = 0  # 結果欄は2列になっているため、分けて書く
        for result in results:
            sheet_str, sheet_num, good_count = result
            # sheet_str = 上　中　下　無の4通りで無は番号のみという意味
            if sheet_str == "無":
                sheet_str = ""
            # シート番号0-20が結果欄の1列目で21-40が2行目
            if sheet_num > 20:
                sheet_num = sheet_num - 20
                column = 3
            else:
                column = 0

            ws.cell(row=row+sheet_num, column=first_column+column).value = sheet_str
            ws.cell(row=row+sheet_num, column=first_column+column+2).value = good_count

        wb.save(save_path)  # 結果の保存
        return True
    except:
        return False
